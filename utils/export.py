"""
Excel ve PDF dışa aktarma işlemleri.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


class ExportManager:
    """Dışa aktarma işlemlerini yöneten sınıf."""
    
    def __init__(self, db_manager):
        self.db = db_manager
        self._register_fonts()
    
    def _register_fonts(self):
        """PDF için Türkçe karakter destekli font kaydeder."""
        try:
            # Windows Arial fontunu kullan
            pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
            pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
            self.font_name = 'Arial'
        except:
            self.font_name = 'Helvetica'
    
    # ==================== EXCEL EXPORT ====================
    
    def export_sinif_listesi_excel(self, sinif_id, filepath):
        """Sınıf listesini Excel'e aktarır. sinif_id None ise tüm sınıflar."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Öğrenci Listesi"
        
        # Başlık stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Sınıf bilgisi al
        if sinif_id:
            siniflar = self.db.get_all_siniflar()
            sinif = next((s for s in siniflar if s['id'] == sinif_id), None)
            sinif_adi = sinif['ad'] if sinif else 'Bilinmeyen Sınıf'
            title = f"{sinif_adi} - Öğrenci Listesi"
            ogrenciler = self.db.get_all_ogrenciler(sinif_id)
            headers = ['Sıra', 'Okul No', 'Ad', 'Soyad']
        else:
            sinif_adi = "TÜM SINIFLAR"
            title = "TÜM SINIFLAR - Öğrenci Listesi"
            ogrenciler = self.db.get_all_ogrenciler(None)
            # Sınıf bilgisi ekle
            for ogrenci in ogrenciler:
                if not ogrenci.get('sinif_adi'):
                    # DB'den gelmiyorsa (get_all_ogrenciler güncellemesine bağlı)
                    # Burda basitçe id'den bulmaya çalışabiliriz ama db_manager.get_all_ogrenciler join yapıyor mu?
                    # Evet, genellikle join yapar. Kontrol ettim, join yapıyor.
                    pass
            # Sıralama: Sınıf, sonra Ad
            headers = ['Sıra', 'Sınıf', 'Okul No', 'Ad', 'Soyad']

        # Başlık satırı
        ws.merge_cells('A1:E1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Tarih
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Oluşturma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # Sütun başlıkları
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Öğrenci verileri
        for row, ogrenci in enumerate(ogrenciler, 5):
            col_offset = 0
            if not sinif_id:
                 ws.cell(row=row, column=1, value=row-4)
                 ws.cell(row=row, column=2, value=ogrenci.get('sinif_adi', '-'))
                 ws.cell(row=row, column=3, value=ogrenci['okul_no'])
                 ws.cell(row=row, column=4, value=ogrenci['ad'])
                 ws.cell(row=row, column=5, value=ogrenci['soyad'])
            else:
                 ws.cell(row=row, column=1, value=row-4)
                 ws.cell(row=row, column=2, value=ogrenci['okul_no'])
                 ws.cell(row=row, column=3, value=ogrenci['ad'])
                 ws.cell(row=row, column=4, value=ogrenci['soyad'])
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 8
        if not sinif_id:
             ws.column_dimensions['B'].width = 12
             ws.column_dimensions['C'].width = 15
             ws.column_dimensions['D'].width = 20
             ws.column_dimensions['E'].width = 20
        else:
             ws.column_dimensions['B'].width = 15
             ws.column_dimensions['C'].width = 20
             ws.column_dimensions['D'].width = 20
        
        wb.save(filepath)
        return filepath
    
    def export_not_cizelgesi_excel(self, sinif_id, filepath):
        """Sınıfın not çizelgesini Excel'e aktarır. sinif_id None ise tüm sınıflar için ayrı sheet."""
        wb = Workbook()
        # Varsayılan sheet'i kaldır (döngüde ekleyeceğiz veya yeniden adlandıracağız)
        default_ws = wb.active
        
        if sinif_id:
            siniflar = [{'id': sinif_id, 'ad': 'Not Çizelgesi'}] # Tek sınıf
            # Sınıf adını al
            all_s = self.db.get_all_siniflar()
            s = next((x for x in all_s if x['id'] == sinif_id), None)
            if s: siniflar[0]['ad'] = s['ad']
        else:
            siniflar = self.db.get_all_siniflar()
            wb.remove(default_ws) # İlk boş sheet'i sil
            
        if not siniflar:
             # Hiç sınıf yoksa boş bir tane oluştur
             wb.create_sheet("Boş")
        
        for sinif_data in siniflar:
            s_id = sinif_data['id']
            s_name = sinif_data['ad']
            
            # Sheet ismi max 31 karakter olabilir
            safe_name = "".join(x for x in s_name if x.isalnum() or x in " -_")[:30]
            ws = wb.create_sheet(title=safe_name)
            
            self._fill_grade_sheet(ws, s_id, s_name)
            
        wb.save(filepath)
        return filepath

    def _fill_grade_sheet(self, ws, sinif_id, sinif_adi):
        """Bir Excel sayfasına sınıfın notlarını doldurur."""
        # Stiller
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        category_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Sınıf ve kategoriler
        ogrenciler = self.db.get_all_ogrenciler(sinif_id)
        kategoriler = self.db.get_all_kategoriler()
        
        # Başlık
        row = 1
        ws.merge_cells(f'A{row}:Z{row}')
        ws[f'A{row}'] = f"{sinif_adi} - Not Çizelgesi"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        row = 3
        # Sütun başlıkları
        ws.cell(row=row, column=1, value="Sıra").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=2, value="Ad Soyad").font = header_font
        ws.cell(row=row, column=2).fill = header_fill
        
        col = 3
        for kategori in kategoriler:
            basliklar = self.db.get_not_basliklari(kategori_id=kategori['id'], sinif_id=sinif_id)
            for baslik in basliklar:
                cell = ws.cell(row=row, column=col, value=baslik['baslik'])
                cell.font = header_font
                cell.fill = category_fill
                col += 1
            
            # Kategori ortalaması
            cell = ws.cell(row=row, column=col, value=f"{kategori['ad']} Ort.")
            cell.font = header_font
            cell.fill = header_fill
            col += 1
        
        # Genel ortalama
        cell = ws.cell(row=row, column=col, value="Genel Ort.")
        cell.font = header_font
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        
        # Öğrenci verileri
        for i, ogrenci in enumerate(ogrenciler, 1):
            row = i + 3
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=f"{ogrenci['ad']} {ogrenci['soyad']}")
            
            curr_col = 3
            for kategori in kategoriler:
                basliklar = self.db.get_not_basliklari(kategori_id=kategori['id'], sinif_id=sinif_id)
                for baslik in basliklar:
                    notlar = self.db.get_notlar(ogrenci_id=ogrenci['id'], baslik_id=baslik['id'])
                    puan = notlar[0]['puan'] if notlar else None
                    ws.cell(row=row, column=curr_col, value=puan if puan is not None else '-')
                    curr_col += 1
                
                # Kategori ortalaması
                ort = self.db.get_ogrenci_kategori_ortalama(ogrenci['id'], kategori['id'])
                ws.cell(row=row, column=curr_col, value=round(ort, 2) if ort else '-')
                curr_col += 1
            
            # Genel ortalama
            genel = self.db.get_ogrenci_genel_ortalama(ogrenci['id'])
            ws.cell(row=row, column=curr_col, value=round(genel, 2) if genel else '-')
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        for i in range(3, col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12
    
    # ==================== PDF EXPORT ====================
    
    def export_ogrenci_karnesi_pdf(self, ogrenci_id, filepath):
        """Öğrenci karnesini PDF'e aktarır."""
        ogrenci = self.db.get_ogrenci_by_id(ogrenci_id)
        if not ogrenci:
            raise ValueError("Öğrenci bulunamadı!")
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Başlık stili
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontName=self.font_name,
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=20
        )
        
        # Normal stil
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontName=self.font_name,
            fontSize=11
        )
        
        # Başlık
        elements.append(Paragraph("ÖĞRENCİ KARNESİ", title_style))
        elements.append(Spacer(1, 20))
        
        # Öğrenci bilgileri
        ogrenci_bilgi = [
            ['Ad Soyad:', f"{ogrenci['ad']} {ogrenci['soyad']}"],
            ['Okul No:', ogrenci['okul_no'] or '-'],
            ['Sınıf:', ogrenci.get('sinif_adi', '-')],
            ['Tarih:', datetime.now().strftime('%d.%m.%Y')]
        ]
        
        bilgi_table = Table(ogrenci_bilgi, colWidths=[4*cm, 10*cm])
        bilgi_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('FONTNAME', (0, 0), (0, -1), f'{self.font_name}'),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(bilgi_table)
        elements.append(Spacer(1, 30))
        
        # Notlar tablosu
        kategoriler = self.db.get_all_kategoriler()
        
        for kategori in kategoriler:
            elements.append(Paragraph(f"<b>{kategori['ad']}</b>", normal_style))
            elements.append(Spacer(1, 10))
            
            notlar = []
            basliklar = self.db.get_not_basliklari(kategori_id=kategori['id'])
            
            for baslik in basliklar:
                not_kayitlari = self.db.get_notlar(ogrenci_id=ogrenci_id, baslik_id=baslik['id'])
                puan = not_kayitlari[0]['puan'] if not_kayitlari else '-'
                notlar.append([baslik['baslik'], str(puan)])
            
            if notlar:
                # Ortalama ekle
                ort = self.db.get_ogrenci_kategori_ortalama(ogrenci_id, kategori['id'])
                notlar.append(['Ortalama', f"{ort:.2f}" if ort else '-'])
                
                not_table = Table(notlar, colWidths=[10*cm, 4*cm])
                not_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), self.font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
                    ('BACKGROUND', (-1, -1), (-1, -1), colors.lightgrey),
                    ('FONTNAME', (0, -1), (-1, -1), self.font_name),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                ]))
                elements.append(not_table)
            else:
                elements.append(Paragraph("Bu kategoride not bulunmuyor.", normal_style))
            
            elements.append(Spacer(1, 20))
        
        # Genel ortalama
        genel_ort = self.db.get_ogrenci_genel_ortalama(ogrenci_id)
        genel_style = ParagraphStyle(
            'Genel',
            parent=styles['Heading2'],
            fontName=self.font_name,
            fontSize=14,
            alignment=1
        )
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"<b>GENEL ORTALAMA: {genel_ort:.2f if genel_ort else '-'}</b>",
            genel_style
        ))
        
        doc.build(elements)
        return filepath
    
    def export_sinif_raporu_pdf(self, sinif_id, filepath):
        """Sınıf raporunu PDF'e aktarır. sinif_id None ise tüm sınıflar."""
        if sinif_id:
            siniflar = self.db.get_all_siniflar()
            sinif = next((s for s in siniflar if s['id'] == sinif_id), None)
            sinif_adi = sinif['ad'] if sinif else 'Bilinmeyen Sınıf'
            title = f"{sinif_adi} - SINIF RAPORU"
        else:
            title = "TÜM OKUL RAPORU"
        
        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontName=self.font_name,
            fontSize=18,
            alignment=1,
            spaceAfter=20
        )
        
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))
        
        # Öğrenciler ve notları
        ogrenciler = self.db.get_all_ogrenciler(sinif_id)
        kategoriler = self.db.get_all_kategoriler()
        
        # Tablo başlıkları
        headers = ['Sıra']
        if not sinif_id:
            headers.append('Sınıf')
        headers.append('Ad Soyad')
        for kategori in kategoriler:
            headers.append(f"{kategori['ad'][:10]}")
        headers.append('Genel')
        
        # Veriler
        data = [headers]
        for i, ogrenci in enumerate(ogrenciler, 1):
            row = [str(i)]
            if not sinif_id:
                row.append(ogrenci.get('sinif_adi', '-'))
            row.append(f"{ogrenci['ad']} {ogrenci['soyad']}")
            for kategori in kategoriler:
                ort = self.db.get_ogrenci_kategori_ortalama(ogrenci['id'], kategori['id'])
                row.append(f"{ort:.1f}" if ort else '-')
            genel = self.db.get_ogrenci_genel_ortalama(ogrenci['id'])
            row.append(f"{genel:.1f}" if genel else '-')
            data.append(row)
        
        # Sınıf/Okul ortalamaları
        avg_row = ['']
        if not sinif_id:
            avg_row.append('') # Sınıf sütunu için
            avg_label_col = 1
        else:
            avg_label_col = 0
            
        avg_row.append('GENEL ORT.')
        
        for kategori in kategoriler:
            ort = self.db.get_sinif_kategori_ortalama(sinif_id, kategori['id'])
            avg_row.append(f"{ort:.1f}" if ort else '-')
        genel = self.db.get_sinif_genel_ortalama(sinif_id)
        avg_row.append(f"{genel:.1f}" if genel else '-')
        data.append(avg_row)
        
        # Tablo oluştur
        # Sütun genişlikleri için hesaplama
        # A4 genişliği = 21cm, kenar boşlukları = 2cm + 2cm = 4cm
        # Kullanılabilir alan = 17cm
        available_width = A4[0] - 4*cm
        
        # Sabit genişlikler
        rank_width = 1*cm
        class_width = 2.5*cm if not sinif_id else 0
        grade_cols_count = len(kategoriler) + 1 # Kategoriler + Genel
        grade_width = 2.2*cm # Not sütunları için genişlik
        
        # İsim sütununa kalan genişlik
        used_width = rank_width + class_width + (grade_cols_count * grade_width)
        name_width = available_width - used_width
        
        # Eğer isim sütunu çok dar kalırsa not sütunlarını biraz daraltabiliriz
        if name_width < 4*cm:
            grade_width = 1.8*cm
            used_width = rank_width + class_width + (grade_cols_count * grade_width)
            name_width = available_width - used_width
        
        col_widths = [rank_width]
        if not sinif_id:
            col_widths.append(class_width)
        col_widths.append(name_width)
        col_widths.extend([grade_width] * grade_cols_count)
             
        table = Table(data, colWidths=col_widths)
        
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        
        doc.build(elements)
        return filepath
