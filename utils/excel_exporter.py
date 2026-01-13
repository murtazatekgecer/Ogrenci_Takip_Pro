"""
Excel export modülü - Sınıf bazlı öğrenci raporları
Görev detayları ile
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, List
import os


def export_to_excel(report_data: Dict[str, List[dict]], categories: List[dict], filepath: str) -> str:
    """
    Öğrenci verilerini Excel dosyasına export et
    
    Args:
        report_data: Sınıf bazlı öğrenci verileri
        categories: Kategori listesi
        filepath: Kaydedilecek dosya yolu
    
    Returns:
        Oluşturulan dosya yolu
    """
    wb = Workbook()
    
    # Varsayılan boş sheet'i sil
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Stiller
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    subheader_fill = PatternFill(start_color='8EA9DB', end_color='8EA9DB', fill_type='solid')
    category_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Koşullu renklendirme için renkler
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    for sinif, students in report_data.items():
        if not students:
            continue
            
        # Her sınıf için yeni sheet
        ws = wb.create_sheet(title=sinif[:31])  # Excel sheet adı max 31 karakter
        
        # Tüm görev isimlerini kategorilere göre topla
        category_tasks = {}  # {kategori_isim: [görev_isimleri]}
        for cat in categories:
            category_tasks[cat['isim']] = set()
            for student in students:
                cat_data = student['kategoriler'].get(cat['isim'], {})
                gorevler = cat_data.get('gorevler', [])
                for gorev in gorevler:
                    category_tasks[cat['isim']].add(gorev['isim'])
            category_tasks[cat['isim']] = sorted(category_tasks[cat['isim']])
        
        # Başlık satırları oluştur
        # Satır 1: Kategori başlıkları (merge cells)
        # Satır 2: Görev isimleri ve ortalama
        
        row1 = 1
        row2 = 2
        current_col = 1
        
        # Sabit sütunlar
        fixed_headers = ['No', 'Numara', 'Ad', 'Soyad']
        for header in fixed_headers:
            ws.merge_cells(start_row=row1, start_column=current_col, end_row=row2, end_column=current_col)
            cell = ws.cell(row=row1, column=current_col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.cell(row=row2, column=current_col).border = thin_border
            current_col += 1
        
        col_widths = [5, 12, 15, 15]  # Sabit sütun genişlikleri
        
        # Kategori sütunları
        category_start_cols = {}
        for cat in categories:
            cat_name = cat['isim']
            tasks = category_tasks.get(cat_name, [])
            
            if not tasks:
                # Görev yoksa sadece ortalama sütunu
                ws.merge_cells(start_row=row1, start_column=current_col, end_row=row1, end_column=current_col)
                cell = ws.cell(row=row1, column=current_col, value=cat_name)
                cell.font = header_font
                cell.fill = category_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                cell2 = ws.cell(row=row2, column=current_col, value='Ort.')
                cell2.font = Font(bold=True)
                cell2.fill = subheader_fill
                cell2.alignment = header_alignment
                cell2.border = thin_border
                
                category_start_cols[cat_name] = {'start': current_col, 'tasks': [], 'avg_col': current_col}
                col_widths.append(10)
                current_col += 1
            else:
                # Görevler + ortalama
                total_cols = len(tasks) + 1  # görevler + ortalama
                category_start_cols[cat_name] = {
                    'start': current_col, 
                    'tasks': tasks, 
                    'avg_col': current_col + len(tasks)
                }
                
                # Kategori başlığını birleştir
                ws.merge_cells(start_row=row1, start_column=current_col, 
                              end_row=row1, end_column=current_col + total_cols - 1)
                cell = ws.cell(row=row1, column=current_col, value=cat_name)
                cell.font = header_font
                cell.fill = category_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                
                # Görev başlıkları
                for i, task_name in enumerate(tasks):
                    cell = ws.cell(row=row2, column=current_col + i, value=task_name[:15])
                    cell.font = Font(bold=True, size=9)
                    cell.fill = subheader_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    col_widths.append(12)
                
                # Ortalama başlığı
                cell = ws.cell(row=row2, column=current_col + len(tasks), value='Ort.')
                cell.font = Font(bold=True)
                cell.fill = subheader_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                col_widths.append(10)
                
                current_col += total_cols
        
        # Genel ortalama sütunu
        ws.merge_cells(start_row=row1, start_column=current_col, end_row=row2, end_column=current_col)
        cell = ws.cell(row=row1, column=current_col, value='Genel\nOrtalama')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.cell(row=row2, column=current_col).border = thin_border
        col_widths.append(12)
        genel_col = current_col
        
        # Öğrenci verilerini yaz
        for row_idx, student in enumerate(students, 3):  # 3. satırdan başla
            col = 1
            
            # Sabit veriler
            ws.cell(row=row_idx, column=col, value=row_idx - 2).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = cell_alignment
            col += 1
            
            ws.cell(row=row_idx, column=col, value=student['numara']).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = cell_alignment
            col += 1
            
            ws.cell(row=row_idx, column=col, value=student['ad']).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = cell_alignment
            col += 1
            
            ws.cell(row=row_idx, column=col, value=student['soyad']).border = thin_border
            ws.cell(row=row_idx, column=col).alignment = cell_alignment
            col += 1
            
            # Kategori verileri
            for cat in categories:
                cat_name = cat['isim']
                cat_info = category_start_cols[cat_name]
                cat_data = student['kategoriler'].get(cat_name, {})
                gorevler = cat_data.get('gorevler', [])
                ortalama = cat_data.get('ortalama', 0)
                
                # Görev puanları için dict oluştur
                task_scores = {g['isim']: g['puan'] for g in gorevler}
                
                # Görev puanlarını yaz
                for i, task_name in enumerate(cat_info['tasks']):
                    puan = task_scores.get(task_name, '')
                    cell = ws.cell(row=row_idx, column=cat_info['start'] + i, value=puan if puan else '')
                    cell.border = thin_border
                    cell.alignment = cell_alignment
                    
                    if puan:
                        if puan >= 70:
                            cell.fill = green_fill
                        elif puan >= 50:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                
                # Ortalama
                cell = ws.cell(row=row_idx, column=cat_info['avg_col'], value=ortalama if ortalama else '')
                cell.border = thin_border
                cell.alignment = cell_alignment
                cell.font = Font(bold=True)
                
                if ortalama:
                    if ortalama >= 70:
                        cell.fill = green_fill
                    elif ortalama >= 50:
                        cell.fill = yellow_fill
                    elif ortalama > 0:
                        cell.fill = red_fill
                
                col = cat_info['avg_col'] + 1
            
            # Genel ortalama
            genel_ort = student.get('genel_ortalama', 0)
            cell = ws.cell(row=row_idx, column=genel_col, value=genel_ort if genel_ort else '')
            cell.border = thin_border
            cell.alignment = cell_alignment
            cell.font = Font(bold=True)
            
            if genel_ort:
                if genel_ort >= 70:
                    cell.fill = green_fill
                elif genel_ort >= 50:
                    cell.fill = yellow_fill
                elif genel_ort > 0:
                    cell.fill = red_fill
        
        # Sütun genişliklerini ayarla
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Dosyayı kaydet
    wb.save(filepath)
    return filepath
